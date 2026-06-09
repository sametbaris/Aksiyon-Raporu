import streamlit as st
import pandas as pd
import re
import requests
import base64
import os
import io
import openpyxl
import gspread
import uuid
from google.oauth2.service_account import Credentials
from datetime import datetime, timedelta
import streamlit.components.v1 as components
from streamlit_autorefresh import st_autorefresh

# ================= SAYFA AYARLARI =================
icon_path = os.path.join("logos", "sistem.png")
st.set_page_config(
    page_title="Aksiyon Raporu", 
    page_icon=icon_path if os.path.exists(icon_path) else "⚖️", 
    layout="wide"
)

# ================= SESSION STATE =================
if "current_view" not in st.session_state:
    st.session_state.current_view = "ana_sayfa"
if "bbx_authenticated" not in st.session_state:
    st.session_state.bbx_authenticated = False

SHEET_ID = "17zVRiwyUYaaEAqyzNx0u7aMMncdgH81vrKbzqS9MHB4"
MAPPING_FILE = "Aksiyon_Mapping_Resimli.xlsx"

# --- PLATFORM ANA LİNKLERİ ---
PLATFORM_LINKS = {
    "Aksiyon": "https://www.akakce.com/braun.html",
    "Braun Shop": "https://www.braunshop.com.tr",
    "Media Markt": "https://www.mediamarkt.com.tr/tr/category/kisisel-bakim-465820.html?brand=ORAL%20B%20OR%20BRAUN%20OR%20REVLON&marketplace=MediaMarkt&sort=availability+asc",
    "Teknosa": "https://www.teknosa.com/kisisel-bakim-c-118?s=%3Arelevance%3Aseller%3Ateknosa%3Abrand%3A2734%3Abrand%3A275%3Abrand%3A2426&text=",
    "Vatan": "https://www.vatanbilgisayar.com/oral-b-braun-revlon/kisisel-bakim-urunleri/?srt=PU",
    "Amazon": "https://www.amazon.com.tr/s?k=braun&rh=p_6%3AA1UNQM1SR2CHM%2Cp_123%3A232435%257C829410&dc&__mk_tr_TR=%C3%85M%C3%85%C5%BD%C3%95%C3%91&crid=1Q148JNG62WA8&qid=1779283345&rnid=91049075031&sprefix=braun%2Caps%2C168&xpid=ykp5xWOeOLnZ9&ref=sr_pg_1",
    "Hepsiburada": "https://www.hepsiburada.com/magaza/hepsiburada?markalar=braun-revlon&kategori=60001547&tab=allproducts",
    "Trendyol": "https://www.trendyol.com/sr?wb=633%2C888&os=1&mid=968"
}

# ================= ŞİFRE POP-UP (MODAL) FONKSİYONU =================
@st.dialog("🔐 BBX Yönetici Girişi")
def login_dialog():
    st.info("ℹ️ Bu alana erişim kısıtlanmıştır. Lütfen yönetici şifresini giriniz.")
    with st.form("login_form"):
        pwd = st.text_input("Şifre:", type="password", key="modal_pwd")
        submitted = st.form_submit_button("Giriş Yap", use_container_width=True)
        if submitted:
            if pwd == "mobese":
                st.session_state.bbx_authenticated = True
                st.session_state.current_view = "bbx_paneli"
                st.rerun()
            else:
                st.error("❌ Hatalı şifre!")

# ================= AKILLI LOGO YÜKLEME =================
def get_base64_logo(file_name):
    file_path = os.path.join("logos", file_name)
    if os.path.exists(file_path):
        try:
            with open(file_path, "rb") as f:
                return f"data:image/png;base64,{base64.b64encode(f.read()).decode()}"
        except: return None
    return None

def load_logo_pair(file_name):
    base_name = file_name.split('.')[0]
    ext = file_name.split('.')[1]
    l_logo = get_base64_logo(file_name)
    d_logo = get_base64_logo(f"{base_name}_white.{ext}")
    return {"light": l_logo, "dark": d_logo if d_logo else l_logo, "invert_dark": not d_logo}

SYSTEM_LOGO = load_logo_pair("sistem.png")

LOGOS = {
    "Aksiyon": load_logo_pair("akakce.png"),
    "Media Markt": load_logo_pair("mediamarkt.png"),
    "Teknosa": load_logo_pair("teknosa.png"),
    "Vatan": load_logo_pair("vatan.png"),
    "Trendyol": load_logo_pair("trendyol.png"),
    "Hepsiburada": load_logo_pair("hepsiburada.png"),
    "Amazon": load_logo_pair("amazon.png"),
    "Braun Shop": load_logo_pair("braunshop.png")
}

# ================= TEMA DEDEKTÖRÜ (GÖLGE DEĞİŞKENLERİ) =================
components.html(
    """
    <script>
    try {
        const parentDoc = window.parent.document;
        if (!parentDoc.getElementById("ninja-referer")) {
            let meta = parentDoc.createElement("meta");
            meta.id = "ninja-referer";
            meta.name = "referrer";
            meta.content = "no-referrer";
            parentDoc.head.appendChild(meta);
        }
        setInterval(() => {
            const bgColor = window.getComputedStyle(parentDoc.body).backgroundColor;
            let rgb = bgColor.match(/\\d+/g);
            if (rgb && rgb.length >= 3) {
                let brightness = (parseInt(rgb[0]) * 299 + parseInt(rgb[1]) * 587 + parseInt(rgb[2]) * 114) / 1000;
                parentDoc.documentElement.style.setProperty('--dynamic-bg-color', bgColor);
                parentDoc.documentElement.style.setProperty('--dynamic-shadow', brightness < 128 ? 'rgba(0, 0, 0, 0.8)' : 'rgba(0, 0, 0, 0.12)');
                let styleTag = parentDoc.getElementById("logo-theme-style");
                if (!styleTag) {
                    styleTag = parentDoc.createElement("style");
                    styleTag.id = "logo-theme-style";
                    parentDoc.head.appendChild(styleTag);
                }
                if (brightness < 128) {
                    styleTag.innerHTML = `.logo-light { display: none !important; } .logo-dark { display: inline-block !important; } .logo-dark.invert-logo { filter: brightness(0) invert(1) !important; } .logo-dark.invert-logo:hover { filter: brightness(0) invert(1) drop-shadow(0px 6px 10px rgba(255,255,255,0.4)) !important; }`;
                    parentDoc.documentElement.classList.add('dark-theme');
                    parentDoc.documentElement.classList.remove('light-theme');
                } else {
                    styleTag.innerHTML = `.logo-dark { display: none !important; } .logo-light { display: inline-block !important; }`;
                    parentDoc.documentElement.classList.add('light-theme');
                    parentDoc.documentElement.classList.remove('dark-theme');
                }
            }
        }, 500);
    } catch (e) {}
    </script>
    """, height=0, width=0
)

# ================= CSS =================
st.markdown("""
<style>
    html, body, [class*="css"] { font-size: 14px !important; }
    .block-container { max-width: 100% !important; padding-top: 1.5rem !important; padding-bottom: 1rem !important; }
    header[data-testid="stHeader"] { height: 2.5rem !important; background: transparent !important; }
    * { -webkit-font-smoothing: antialiased !important; -moz-osx-font-smoothing: grayscale !important; text-rendering: optimizeLegibility !important; }
    :root { --header-color: #888; --pill-default-bg: rgba(128, 128, 128, 0.1); }
    
    .logo-light { display: inline-block !important; }
    .logo-dark { display: none !important; }
    .dark-theme .logo-light, html[data-theme="dark"] .logo-light { display: none !important; }
    .dark-theme .logo-dark, html[data-theme="dark"] .logo-dark { display: inline-block !important; }
    .dark-theme .logo-dark.invert-logo, html[data-theme="dark"] .logo-dark.invert-logo { filter: brightness(0) invert(1) !important; }

    .main-logo-container { display: flex; align-items: center; gap: 15px; margin-bottom: 20px; flex-wrap: wrap; }
    .main-system-logo { height: 50px; width: auto; object-fit: contain; transition: height 0.3s; }
    .main-title-text { margin: 0; display: inline-block; font-size: 1.8rem; font-weight: 700; transition: font-size 0.3s; }
    .online-badge-container { display: flex; align-items: center; gap: 6px; background: rgba(0, 255, 0, 0.1); padding: 4px 12px; border-radius: 20px; border: 1px solid rgba(0, 255, 0, 0.2); }
    
    @media (max-width: 768px) {
        .main-logo-container { flex-direction: column; align-items: center; justify-content: center; text-align: center; gap: 8px; width: 100%; }
        .main-system-logo { height: 40px; }
        .main-title-text { font-size: 1.4rem; }
        .online-badge-container { margin-left: 0 !important; margin-top: 2px; }
        .update-badge { float: none !important; margin: 0 auto 15px auto !important; width: fit-content !important; display: block !important; }
    }
    
    div[data-baseweb="popover"] { transition: none !important; animation: none !important; will-change: auto !important; }
    div[data-baseweb="popover"] ul { transform: translateZ(0) !important; backface-visibility: hidden !important; }
    div[data-baseweb="popover"] [role="option"], div[data-baseweb="popover"] [role="option"] span { text-rendering: optimizeLegibility !important; font-family: inherit !important; font-size: 13px !important; font-weight: 500 !important; letter-spacing: normal !important; }
    
    .table-container { width: 100%; margin-top: 10px; overflow: auto; max-height: 65vh; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); border: none !important; }
    .table-container::-webkit-scrollbar { width: 5px !important; height: 5px !important; }
    .table-container::-webkit-scrollbar-track { background: transparent !important; }
    .table-container::-webkit-scrollbar-thumb { background-color: rgba(128, 128, 128, 0) !important; border-radius: 10px !important; transition: background-color 0.3s ease-in-out !important; }
    .table-container:hover::-webkit-scrollbar-thumb { background-color: rgba(128, 128, 128, 0.15) !important; }
    .table-container::-webkit-scrollbar-thumb:hover { background-color: rgba(128, 128, 128, 0.20) !important; }
    ::-webkit-scrollbar-button, *::-webkit-scrollbar-button, ::-webkit-scrollbar-button:vertical, ::-webkit-scrollbar-button:horizontal, ::-webkit-scrollbar-button:start, ::-webkit-scrollbar-button:end, ::-webkit-scrollbar-button:decrement, ::-webkit-scrollbar-button:increment { display: none !important; width: 0px !important; height: 0px !important; size: 0px !important; background: transparent !important; border: none !important; }
    
    .table-container { scrollbar-width: thin !important; scrollbar-color: rgba(128, 128, 128, 0) transparent !important; transition: scrollbar-color 0.3s ease-in-out !important; }
    .table-container:hover { scrollbar-color: rgba(128, 128, 128, 0.15) transparent !important; }
    
    .custom-table { width: 100%; table-layout: auto; border-collapse: separate !important; border-spacing: 0 !important; font-family: 'Inter', sans-serif; border: none !important; }
    
    .header-logo { height: 26px; width: auto; max-width: 120px; object-fit: contain; transition: transform 0.3s cubic-bezier(0.25, 0.8, 0.25, 1), filter 0.3s ease; will-change: transform, filter; }
    .header-logo:hover { transform: scale3d(1.15, 1.15, 1) translateZ(0); filter: drop-shadow(0px 3px 5px rgba(0,0,0,0.25)); }
    .dark-theme .logo-dark.invert-logo:hover, html[data-theme="dark"] .logo-dark.invert-logo:hover { filter: brightness(0) invert(1) drop-shadow(0px 6px 10px rgba(255,255,255,0.4)) !important; }
    
    .custom-table thead th { position: sticky; top: 0px !important; z-index: 50 !important; padding: 12px 18px; text-align: center; color: var(--header-color); font-weight: 500; text-transform: uppercase; font-size: 10px; background-color: var(--dynamic-bg-color, #ffffff) !important; box-shadow: 0 -2px 0 var(--dynamic-bg-color, #ffffff), 0 8px 15px -4px var(--dynamic-shadow, rgba(0,0,0,0.15)) !important; border-top: none !important; border-left: none !important; border-right: none !important; border-bottom: 1px solid rgba(128,128,128,0.1) !important; }
    .custom-table td { padding: 8px 10px; text-align: center; white-space: nowrap; border-top: none !important; border-left: none !important; border-right: none !important; border-bottom: 1px solid rgba(128,128,128,0.06) !important; }
    .custom-table tbody tr:last-child td { border-bottom: none !important; }
    
    .data-link { text-decoration: none; color: inherit; display: inline-block; width: 100%; }
    .data-pill { padding: 5px 12px; display: inline-flex; align-items: center; justify-content: center; border-radius: 20px; font-size: 13px; line-height: 1.2; transform: translateY(0); transition: transform 0.3s cubic-bezier(0.25, 0.8, 0.25, 1), box-shadow 0.3s cubic-bezier(0.25, 0.8, 0.25, 1); backface-visibility: hidden; -webkit-font-smoothing: antialiased; will-change: transform; }
    a.data-link:hover .data-pill { transform: translateY(-3px); box-shadow: 0px 5px 12px rgba(0,0,0,0.15); cursor: pointer; }

    /* THUMBNAIL GÖRSEL CSS (HEM ANA SAYFA HEM BBX İÇİN ORTAK KULLANIM) */
    .sku-wrapper { position: relative; display: inline-block; cursor: pointer; }
    .sku-thumb { visibility: hidden; position: absolute; left: 110%; top: 50%; transform: translateY(-50%) translateX(10px); opacity: 0; transition: all 0.2s ease-in-out; background-color: var(--dynamic-bg-color, #ffffff); padding: 5px; border-radius: 12px; box-shadow: 0 10px 40px rgba(0,0,0,0.3); z-index: 999999 !important; border: 1px solid rgba(128,128,128,0.2); pointer-events: none; width: 162px !important; height: 162px !important; display: flex !important; align-items: center !important; justify-content: center !important; }
    .sku-thumb img { width: 150px !important; height: 150px !important; min-width: 150px !important; min-height: 150px !important; max-width: 150px !important; object-fit: contain !important; border-radius: 8px; background: white; display: block !important; }
    .sku-thumb::after { content: ''; position: absolute; top: 50%; right: 100%; margin-top: -8px; border-width: 8px; border-style: solid; border-color: transparent var(--dynamic-bg-color, #ffffff) transparent transparent; }
    .sku-wrapper:hover .sku-thumb { visibility: visible; opacity: 1; transform: translateY(-50%) translateX(0px); }

    .update-badge { text-align: right; color: var(--header-color); font-size: 11px; background: var(--pill-default-bg); padding: 5px 14px; border-radius: 30px; display: inline-block; float: right; margin-top: 10px; }
    div[data-testid="stDownloadButton"] button, div[data-testid="stButton"] button { width: 100%; border-radius: 20px; font-weight: 600; border: 1px solid #ddd; font-size: 13px; padding: 4px 8px; }
    div[data-testid="stDownloadButton"] button p, div[data-testid="stButton"] button p { display: flex; align-items: center; justify-content: center; text-align: center; white-space: normal; line-height: 1.2; margin: 0; height: 100%; }
    
    @media (max-width: 950px) and (orientation: landscape) {
        div[data-testid="stButton"] button p, div[data-testid="stDownloadButton"] button p { font-size: 0px !important; }
        div[data-testid="stButton"] button p::before { content: "🧹"; font-size: 16px !important; visibility: visible; }
        div[data-testid="stDownloadButton"] button p::before { content: "📥"; font-size: 16px !important; visibility: visible; }
    }
    .logo-dark { display: none; }
</style>
""", unsafe_allow_html=True)

# ================= GSPREAD KİMLİK DOĞRULAMA =================
@st.cache_resource
def get_gspread_client():
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            return gspread.service_account_from_dict(creds_dict)
        elif os.path.exists("service_account.json"):
            return gspread.service_account(filename="service_account.json")
        else: return None
    except Exception as e:
        print(f"Auth Error: {e}")
        return None

# ================= BBX İÇİN VERİ ÇEKME VE GÖRSEL MAPPING =================
def get_bbx_data_from_sheets():
    try:
        client = get_gspread_client()
        if not client:
            st.error("Google Sheets istemcisi oluşturulamadı.")
            return []
        return client.open_by_key(SHEET_ID).worksheet("BBX").get_all_values()
    except Exception as e:
        st.error(f"Google Sheets BBX verisi çekilirken hata: {e}")
        return []

@st.cache_data(ttl=600)
def get_image_mapping():
    img_map = {}
    if os.path.exists(MAPPING_FILE):
        try:
            df = pd.read_excel(MAPPING_FILE, engine='openpyxl', dtype=str)
            bc_col = next((c for c in df.columns if "barkod" in c.lower()), None)
            if bc_col and "Gorsel_URL" in df.columns:
                for _, row in df.iterrows():
                    bc = clean_val(row[bc_col])
                    img = str(row["Gorsel_URL"]).strip()
                    if bc and img.startswith("http"):
                        img_map[bc] = img
        except: pass
    return img_map

# ================= ZİYARETÇİ TAKİP =================
@st.cache_data(ttl=60)
def get_online_count():
    client = get_gspread_client()
    if not client: return 1
    try:
        sh = client.open_by_key(SHEET_ID)
        log_sheet = sh.worksheet("Ziyaretci_Log")
        all_logs = log_sheet.get_all_records()
        online_count = 0
        tr_now = datetime.utcnow() + timedelta(hours=3)
        two_minutes_ago = tr_now - timedelta(minutes=2)
        for record in all_logs:
            try:
                last_seen = datetime.strptime(str(record.get('Son_Gorulme', '')), "%Y-%m-%d %H:%M:%S")
                if last_seen > two_minutes_ago: online_count += 1
            except: pass
        return max(1, online_count)
    except: return 1

def track_user_presence():
    if 'user_id' not in st.session_state:
        st.session_state.user_id = str(uuid.uuid4())[:8]
        st.session_state.last_ping = None
    now = datetime.utcnow() + timedelta(hours=3)
    client = get_gspread_client()
    if client and (st.session_state.last_ping is None or (now - st.session_state.last_ping).total_seconds() > 60):
        try:
            sh = client.open_by_key(SHEET_ID)
            log_sheet = sh.worksheet("Ziyaretci_Log")
            now_str = now.strftime("%Y-%m-%d %H:%M:%S")
            cell = log_sheet.find(st.session_state.user_id)
            if cell: log_sheet.update_cell(cell.row, 2, now_str)
            else: log_sheet.append_row([st.session_state.user_id, now_str])
            st.session_state.last_ping = now
        except: pass
    return get_online_count()

# ================= YARDIMCI FONKSİYONLAR =================
def clean_val(val):
    if pd.isna(val) or str(val).strip().lower() in ["nan", "none", ""]: return ""
    v = str(val).strip()
    if v.startswith("http"): return v
    return v.split('.')[0]

def parse_price(val):
    if not val or pd.isna(val) or str(val).lower() in ["nan", "none", ""]: return None
    val_str = str(val).lower().replace("tl", "").replace("₺", "").replace(".", "").replace(",", ".").strip()
    clean = re.sub(r"[^\d.]", "", val_str)
    try: return float(clean)
    except: return None

def get_column_mapping(df):
    def find_col(name_part, exclude=None):
        for c in df.columns:
            if name_part.lower() in c.lower():
                if exclude and exclude.lower() in c.lower(): continue
                return c
        return None
    return {
        "Marka": find_col("Marka"), "Ürün Adı": find_col("Ürün Adı"),
        "Barkod": find_col("Barkod"), "Ürün Kodu": find_col("Kodu", exclude="Barkod"),
        "Alt Grup": find_col("Grup"), "Aksiyon": find_col("Aksiyon") or find_col("Akakçe"),
        "Braun Shop": find_col("Braun Shop"), "Media Markt": find_col("Media Markt"),
        "Teknosa": find_col("Teknosa"), "Vatan": find_col("Vatan"),
        "Trendyol": find_col("Trendyol"), "Hepsiburada": find_col("Hepsiburada") or find_col("Hepsi"),
        "Amazon": find_col("Amazon")
    }

# ================= AKILLI LİNK MOTORU =================
def build_smart_link(label, raw_id, row):
    sheet_url = str(row.get(f"{label}_URL", "")).strip()
    if sheet_url.startswith("http"): 
        return sheet_url
    elif sheet_url.startswith("/"):
        if label == "Aksiyon": return f"https://www.akakce.com{sheet_url}"
        if label == "Braun Shop": return f"https://www.braunshop.com.tr{sheet_url}"
        if label == "Media Markt": return f"https://www.mediamarkt.com.tr{sheet_url}"
        if label == "Teknosa": return f"https://www.teknosa.com{sheet_url}"
        if label == "Vatan": return f"https://www.vatanbilgisayar.com{sheet_url}"
        if label == "Trendyol": return f"https://www.trendyol.com{sheet_url}"
        if label == "Hepsiburada": return f"https://www.hepsiburada.com{sheet_url}"
        
    val = clean_val(raw_id)
    barcode = clean_val(row.get("Barkod_Int", ""))
    
    if label == "Aksiyon":
        if barcode: return f"https://www.akakce.com/arama/?q={barcode}"
        if val: return f"https://www.akakce.com/arama/?q={val}"
        return None
        
    if label == "Braun Shop":
        if val: return f"https://www.braunshop.com.tr/index.php?route=product/product&product_id={val}"
        if barcode: return f"https://www.braunshop.com.tr/arama?q={barcode}"
        return None
        
    if val != "":
        if label == "Trendyol": return f"https://www.trendyol.com/brand/product-p-{val}"
        if label == "Hepsiburada": return f"https://www.hepsiburada.com/product-p-{val}"
        if label == "Amazon": return f"https://www.amazon.com.tr/dp/{val}"
        if label == "Media Markt": return f"https://www.mediamarkt.com.tr/tr/product/_{val}.html"
        
    if barcode:
        if label == "Media Markt": return f"https://www.mediamarkt.com.tr/tr/search.html?query={barcode}"
        if label == "Teknosa": return f"https://www.teknosa.com/arama/?s={barcode}"
        if label == "Vatan": return f"https://www.vatanbilgisayar.com/arama/{barcode}/"
        
    return None

# ================= GİZLİ BAĞLANTI & VERİ BİRLEŞTİRME =================
@st.cache_data(ttl=180)  
def load_and_merge_data():
    client = get_gspread_client()
    if not client:
        st.error("🔒 Güvenlik Hatası: Streamlit Secrets'ta JSON Bilgileri Bulunamadı!")
        return None, ""

    try:
        sh = client.open_by_key(SHEET_ID)
        worksheet = sh.worksheet("Guncel")
        
        update_text = ""
        try:
            update_text = str(worksheet.acell("N1").value).replace('"', '').strip()
        except: pass

        data = worksheet.get_all_values()
        if not data: return None, update_text
        
        try:
            data_formulas = worksheet.get_all_values(value_render_option='FORMULA')
        except:
            data_formulas = data
            
        df_fiyat = pd.DataFrame(data[1:], columns=[str(c).strip() for c in data[0]])
        df_formulas = pd.DataFrame(data_formulas[1:], columns=[str(c).strip() for c in data_formulas[0]])
        
        def extract_url(val):
            if isinstance(val, str) and 'HYPERLINK' in val.upper():
                m = re.search(r'HYPERLINK\(\s*["\']([^"\']+)["\']', val, re.IGNORECASE)
                if m: return m.group(1)
            return ""
            
        br_col = next((c for c in df_formulas.columns if "braun" in c.lower() and "kodu" in c.lower()), None)
        if br_col: df_fiyat["Aksiyon_URL"] = df_formulas[br_col].apply(extract_url)
        else: df_fiyat["Aksiyon_URL"] = ""
        
        for plat in ["Braun Shop", "Media Markt", "Teknosa", "Vatan", "Trendyol", "Hepsiburada", "Amazon"]:
            col = next((c for c in df_formulas.columns if plat.lower().replace(" ", "") in c.lower().replace(" ", "") or (plat=="Hepsiburada" and "hepsi" in c.lower())), None)
            if col: df_fiyat[f"{plat}_URL"] = df_formulas[col].apply(extract_url)
            else: df_fiyat[f"{plat}_URL"] = ""

        bc_col = next((c for c in df_fiyat.columns if "barkod" in c.lower()), None)
        if bc_col: df_fiyat["Barkod_Int"] = df_fiyat[bc_col].apply(clean_val)
        else: df_fiyat["Barkod_Int"] = ""
        
        if os.path.exists(MAPPING_FILE):
            df_map = pd.read_excel(MAPPING_FILE, engine='openpyxl', dtype=str)
            df_map.columns = [c.strip() for c in df_map.columns]
            
            map_bc_col = next((c for c in df_map.columns if "barkod" in c.lower()), None)
            if map_bc_col and map_bc_col in df_map.columns:
                df_map["Barkod_Int"] = df_map[map_bc_col].apply(clean_val)
            else:
                df_map["Barkod_Int"] = ""
                
            link_cols = ["Barkod_Int", "Gorsel_URL", "Marka"]
            df_map_sub = df_map[[c for c in link_cols if c in df_map.columns]].copy()
            df_final = pd.merge(df_fiyat, df_map_sub, on="Barkod_Int", how="left")
            return df_final.fillna(""), update_text
            
        return df_fiyat.fillna(""), update_text
        
    except Exception as e:
        st.error(f"Google bağlantı hatası: {e}")
        return None, ""

# ================= RENDER TABLO =================
def display_styled_table(df, mapping):
    refs = { "Aksiyon": "CSS Code", "Braun Shop": "BS Data ID", "Media Markt": "MM", "Teknosa": "TKNS", "Vatan": "VTN", "Trendyol": "TY", "Hepsiburada": "HB", "Amazon": "AMZ" }
    pazaryerleri = ["Aksiyon", "Braun Shop", "Media Markt", "Teknosa", "Vatan", "Trendyol", "Hepsiburada", "Amazon"]
    
    html = '<div class="table-container"><table class="custom-table"><thead><tr>'
    
    for label, real in mapping.items():
        if label == "Marka": continue 
        if real:
            count_html = ""
            if label in pazaryerleri:
                valid_count = sum(1 for v in df[real] if parse_price(v) is not None)
                count_html = f'<div style="font-size: 10px; color: var(--header-color); margin-top: 5px; font-weight: 700; letter-spacing: 0.5px; text-transform: none;">{valid_count} Ürün</div>'

            logo_pair = LOGOS.get(label)
            plat_url = PLATFORM_LINKS.get(label)
            
            if logo_pair and logo_pair["light"]:
                l_src = logo_pair["light"]; d_src = logo_pair["dark"]
                inv_class = "invert-logo" if logo_pair["invert_dark"] and label in ["Amazon", "Aksiyon"] else ""
                if plat_url: content = f'<a href="{plat_url}" target="_blank" style="text-decoration:none;"><img src="{l_src}" class="header-logo logo-light" title="{label}"><img src="{d_src}" class="header-logo logo-dark {inv_class}" title="{label}"></a>'
                else: content = f'<img src="{l_src}" class="header-logo logo-light" title="{label}"><img src="{d_src}" class="header-logo logo-dark {inv_class}" title="{label}">'
                html += f'<th>{content}{count_html}</th>'
            else: html += f'<th>{label}{count_html}</th>'

    html += '</tr></thead><tbody>'
    
    for _, row in df.iterrows():
        html += '<tr>'
        for label, real in mapping.items():
            if not real or label == "Marka": continue 
            
            val = str(row[real]); d_val = "" if val.lower() in ["nan", "none", ""] else val; style = ""
            bs_col_name = mapping.get("Braun Shop")
            
            if label in ["Media Markt", "Teknosa", "Vatan", "Trendyol", "Hepsiburada", "Amazon"] and bs_col_name:
                p_ref = parse_price(row[bs_col_name]); p_curr = parse_price(d_val)
                if p_ref and p_curr:
                    if p_curr == p_ref: style = 'background-color: #d4edda; color: #155724; font-weight: 600;' 
                    elif p_curr > p_ref: style = 'background-color: #fff3cd; color: #856404; font-weight: 600;' 
                    else: style = 'background-color: #f8d7da; color: #721c24; font-weight: 600;' 
                    
            if not style and d_val:
                if any(x in label.lower() for x in ["barkod", "kodu", "grup", "marka"]): style = 'background-color: transparent;'
                else: style = 'background-color: var(--pill-default-bg);'
                
            map_key = refs.get(label); target_id = row.get(map_key, "")
            url = build_smart_link(label, target_id, row)
            
            img_url = str(row.get("Gorsel_URL", "")).strip()
            is_sku_col = (label == "Ürün Kodu")
            has_img = is_sku_col and img_url.startswith("http")
            
            inner_content = f'<span class="data-pill" style="{style}">{d_val}</span>'
            
            if has_img:
                inner_content = f'<div class="sku-wrapper">{inner_content}<div class="sku-thumb"><img src="{img_url}" referrerpolicy="no-referrer"></div></div>'
            
            if url and d_val: html += f'<td><a href="{url}" target="_blank" class="data-link">{inner_content}</a></td>'
            else: html += f'<td>{inner_content}</td>'
        html += '</tr>'
    st.markdown(html + '</tbody></table></div>', unsafe_allow_html=True)

# ================= SESSION STATE BAŞLATMA (FİLTRELER İÇİN) =================
if "search_val" not in st.session_state: st.session_state.search_val = ""
if "marka_val" not in st.session_state: st.session_state.marka_val = [] 
if "grup_val" not in st.session_state: st.session_state.grup_val = []
if "plat_val" not in st.session_state: st.session_state.plat_val = None
if "stat_val" not in st.session_state: st.session_state.stat_val = None

if "search_bbx" not in st.session_state: st.session_state.search_bbx = ""
if "grup_bbx" not in st.session_state: st.session_state.grup_bbx = []
if "stat_bbx" not in st.session_state: st.session_state.stat_bbx = "Tümü"

def reset_filters():
    st.session_state.search_val = ""
    st.session_state.marka_val = [] 
    st.session_state.grup_val = []
    st.session_state.plat_val = None
    st.session_state.stat_val = None

def reset_bbx_filters():
    st.session_state.search_bbx = ""
    st.session_state.grup_bbx = []
    st.session_state.stat_bbx = "Tümü"


# ==============================================================
# ================= YÖNLENDİRİCİ BLOK BAŞLIYOR =================
# ==============================================================

if st.session_state.current_view == "ana_sayfa":

    # ================= MAIN =================
    col_title, col_update = st.columns([3, 1])

    with col_title:
        online_users = track_user_presence()
        online_badge = f'<div class="online-badge-container"><span style="height: 8px; width: 8px; background-color: #00ff00; border-radius: 50%; display: inline-block; box-shadow: 0 0 8px #00ff00; margin-right: 6px;"></span><span style="color: #00ff00; font-size: 11px; font-weight: 600; white-space: nowrap;">{online_users} Online</span></div>'

        if SYSTEM_LOGO["light"]:
            l_sys = SYSTEM_LOGO["light"]
            st.markdown(f'<div class="main-logo-container"><img src="{l_sys}" class="main-system-logo"><h1 class="main-title-text">Aksiyon Raporu</h1>{online_badge}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="main-logo-container"><h1 class="main-title-text">📊 Aksiyon Raporu</h1>{online_badge}</div>', unsafe_allow_html=True)

    res = load_and_merge_data()
    df_data = res[0] if res else None
    update_text = res[1] if res else ""

    with col_update:
        if update_text: 
            st.markdown(f'''
                <div class="update-badge">🔄 {update_text}</div>
                <div style="clear: both; height: 10px;"></div>
            ''', unsafe_allow_html=True)
        else:
            st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
            
        _, btn_col = st.columns([1, 1.4]) 
        with btn_col:
            if st.button("🔐 BBX Paneli", use_container_width=True):
                if not st.session_state.bbx_authenticated:
                    login_dialog()
                else:
                    st.session_state.current_view = "bbx_paneli"
                    st.rerun()

    if df_data is not None:
        mapping = get_column_mapping(df_data)
        alt_grup_col = mapping.get("Alt Grup")
        marka_col = mapping.get("Marka")
        
        if alt_grup_col and alt_grup_col in df_data.columns:
            gruplar = []
            for x in df_data[alt_grup_col].dropna():
                v = str(x).strip()
                if v != "" and v not in gruplar: gruplar.append(v)
        else: 
            gruplar = []
            
        if marka_col and marka_col in df_data.columns:
            markalar_raw = []
            for x in df_data[marka_col].dropna():
                v = str(x).strip()
                if v != "" and v not in markalar_raw: markalar_raw.append(v)
            preferred_order = ["Braun", "Oral-B", "Braun Saç", "Revlon Saç"]
            markalar = sorted(markalar_raw, key=lambda x: preferred_order.index(x) if x in preferred_order else 999)
        else:
            markalar = []

        col_search, col_marka, col_grup, col_plat, col_stat, col_btn_group = st.columns([1.9, 1.7, 1.7, 1.7, 1.4, 2.0])
        
        with col_search: 
            search = st.text_input("🔍 Ürün Ara...", key="search_val")
        with col_marka:
            filter_marka = st.multiselect("🏷️ Marka", markalar, placeholder="Tümü", key="marka_val")
        with col_grup: 
            filter_grup = st.multiselect("📂 Alt Grup", gruplar, placeholder="Tümü", key="grup_val")
        with col_plat: 
            filter_platform = st.selectbox("🛒 Platform", ["Media Markt", "Teknosa", "Vatan", "Trendyol", "Hepsiburada", "Amazon"], index=None, placeholder="Tümü", key="plat_val")
        with col_stat: 
            filter_status = st.selectbox("🎨 Renge Göre", ["🔴 Kırmızı (↓)", "🟢 Yeşil (=)", "🟡 Sarı (↑)"], index=None, placeholder="Tümü", key="stat_val")

        if search: df_data = df_data[df_data.apply(lambda r: r.astype(str).str.contains(search, case=False).any(), axis=1)]
        
        if filter_marka and marka_col:
            df_data = df_data[df_data[marka_col].astype(str).str.strip().isin(filter_marka)]
            
        if filter_grup and alt_grup_col: 
            if "Tümü" not in filter_grup: 
                df_data = df_data[df_data[alt_grup_col].astype(str).str.strip().isin(filter_grup)]
                
        if filter_status:
            bs_col = mapping.get("Braun Shop")
            if bs_col:
                p_list = [filter_platform] if filter_platform else ["Media Markt", "Teknosa", "Vatan", "Trendyol", "Hepsiburada", "Amazon"]
                actual_cols = [mapping.get(p) for p in p_list]; actual_cols = [c for c in actual_cols if c] 
                def check_color(row):
                    bs_val = parse_price(row[bs_col])
                    if bs_val is None: return False
                    for col in actual_cols:
                        p_val = parse_price(row[col])
                        if p_val is not None:
                            if "🔴" in filter_status and p_val < bs_val: return True
                            if "🟢" in filter_status and p_val == bs_val: return True
                            if "🟡" in filter_status and p_val > bs_val: return True
                    return False
                df_data = df_data[df_data.apply(check_color, axis=1)]

        # ================= EXCEL İNDİRME VE FORMATLAMA =================
        tr_time_now = datetime.utcnow() + timedelta(hours=3)
        current_time_str = tr_time_now.strftime("%d-%m-%Y_%H-%M")
        excel_filename = f"Aksiyon_Raporu_{current_time_str}.xlsx"
        
        export_cols = [real for label, real in mapping.items() if real in df_data.columns and label != "Marka"]
        df_export = df_data[export_cols].copy()
        
        price_platforms = ["Aksiyon", "Braun Shop", "Media Markt", "Teknosa", "Vatan", "Trendyol", "Hepsiburada", "Amazon"]
        price_cols = [mapping.get(p) for p in price_platforms if mapping.get(p) in df_export.columns]
        for col_name in price_cols:
            df_export[col_name] = df_export[col_name].apply(parse_price)
            
        barcode_col = mapping.get("Barkod")
        if barcode_col and barcode_col in df_export.columns:
            def parse_barcode(v):
                try: return int(float(str(v).replace(' ', '')))
                except: return v
            df_export[barcode_col] = df_export[barcode_col].apply(parse_barcode)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Aksiyon_Raporu')
            workbook = writer.book
            worksheet = writer.sheets['Aksiyon_Raporu']
            
            for idx, col_name in enumerate(df_export.columns):
                excel_col_idx = idx + 1
                col_letter = openpyxl.utils.get_column_letter(excel_col_idx)
                
                if col_name in price_cols:
                    worksheet.column_dimensions[col_letter].width = 12
                    for row in range(2, len(df_export) + 2):
                        worksheet.cell(row=row, column=excel_col_idx).number_format = '#,##0.00'
                        
                elif col_name == barcode_col:
                    worksheet.column_dimensions[col_letter].width = 16
                    for row in range(2, len(df_export) + 2):
                        worksheet.cell(row=row, column=excel_col_idx).number_format = '0'
                else:
                    worksheet.column_dimensions[col_letter].width = 15
                    
        with col_btn_group:
            st.markdown("<div style='margin-top: 23px;'></div>", unsafe_allow_html=True)
            btn_clear, btn_excel = st.columns([1, 1])
            with btn_clear:
                st.button("🧹 Filtre Temizle", on_click=reset_filters, use_container_width=True)
            with btn_excel:
                st.download_button("📥 Excel'e Aktar", output.getvalue(), excel_filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        display_styled_table(df_data, mapping)

    st_autorefresh(interval=180000, limit=None, key="data_refresher")


elif st.session_state.current_view == "bbx_paneli":
    # ==============================================================
    # ================= ŞİFRELİ BBX KONTROL PANELİ =================
    # ==============================================================
    
    st.markdown("""
    <style>
        .bbx-title { background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent; font-size: 2.2rem; font-weight: 800; margin-bottom: 0.1rem; font-family: 'Inter', sans-serif; }
        .bbx-subtitle { color: var(--text-color); opacity: 0.7; font-size: 1.0rem; margin-bottom: 0.5rem; font-family: 'Inter', sans-serif; }
        
        .bbx-table-wrapper { 
            background-color: transparent; 
            border-radius: 8px; 
            padding: 0px; 
            overflow: auto; 
            max-height: 65vh; 
            margin-top: 5px; 
            box-shadow: 0 2px 4px rgba(0,0,0,0.05); 
        }
        .bbx-table-wrapper::-webkit-scrollbar { width: 5px !important; height: 5px !important; }
        .bbx-table-wrapper::-webkit-scrollbar-track { background: transparent !important; }
        .bbx-table-wrapper::-webkit-scrollbar-thumb { background-color: rgba(128, 128, 128, 0) !important; border-radius: 10px !important; transition: background-color 0.3s ease-in-out !important; }
        .bbx-table-wrapper:hover::-webkit-scrollbar-thumb { background-color: rgba(128, 128, 128, 0.15) !important; }
        .bbx-table-wrapper::-webkit-scrollbar-thumb:hover { background-color: rgba(128, 128, 128, 0.20) !important; }
        
        .bbx-table { width: 100%; table-layout: auto; border-collapse: separate !important; border-spacing: 0 !important; font-family: 'Inter', sans-serif; font-size: 13px; color: var(--text-color); border: none !important; }
        
        .bbx-table thead th { 
            position: sticky; 
            z-index: 50 !important; 
            padding: 12px 6px; 
            text-align: center; 
            color: var(--header-color); 
            font-weight: 600; 
            text-transform: uppercase; 
            font-size: 11px; 
            background-color: var(--dynamic-bg-color, #ffffff) !important; 
            border-top: none !important; 
            border-left: none !important; 
            border-right: none !important; 
        }
        
        /* GÖLGELİ VE SABİT BAŞLIK SİHİRBAZLIĞI (58px OFFSET - SIFIR BOŞLUK) */
        .bbx-table thead tr:nth-child(1) th { top: 0px !important; box-shadow: none !important; border-bottom: none !important; }
        .bbx-table thead tr:nth-child(2) th { top: 58px !important; box-shadow: 0 8px 15px -4px var(--dynamic-shadow, rgba(0,0,0,0.15)) !important; border-bottom: 1px solid rgba(128,128,128,0.1) !important; border-top: none !important; }
        
        /* DÜZELTME: İlk satırda Barkod, SKU gibi rowspan=2 olan hücrelere gölgeyi geri getiriyoruz */
        .bbx-table thead tr:nth-child(1) th[rowspan="2"] { 
            box-shadow: 0 8px 15px -4px var(--dynamic-shadow, rgba(0,0,0,0.15)) !important; 
            border-bottom: 1px solid rgba(128,128,128,0.1) !important; 
        }
        
        .bbx-table td { 
            border-top: none !important; 
            border-left: none !important; 
            border-right: none !important; 
            border-bottom: 1px solid rgba(128,128,128,0.06) !important; 
            padding: 10px 6px; 
            text-align: center; 
            vertical-align: middle; 
        }
        
        .plat-sep { border-left: 2px solid rgba(128,128,128,0.15) !important; }
        
        .bbx-sku { color: #0078ff !important; font-weight: bold; font-size: 13px; } 
        .pill-satici { background-color: var(--pill-default-bg); border: 1px solid rgba(128, 128, 128, 0.2); border-radius: 20px; padding: 4px 10px; font-weight: 600; font-size: 11px; display: inline-block; max-width: 130px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
        .pill-fiyat { background-color: rgba(0, 184, 76, 0.1); border: 1px solid rgba(0, 184, 76, 0.3); color: #00b84c; border-radius: 20px; padding: 4px 10px; font-weight: bold; font-size: 12px; display: inline-block; }
        
        @keyframes blink { 0% { opacity: 1; transform: scale(1); } 50% { opacity: 0.4; transform: scale(1.1); } 100% { opacity: 1; transform: scale(1); } }
        .status-dot { height: 12px; width: 12px; border-radius: 50%; display: inline-block; animation: blink 1.5s infinite; }
        .dot-green { background-color: #00e676; box-shadow: 0 0 10px #00e676; }
        .dot-yellow { background-color: #ffea00; box-shadow: 0 0 10px #ffea00; }
        .dot-red { background-color: #ff1744; box-shadow: 0 0 10px #ff1744; }
        .n-b-b { border-bottom: none !important; padding-bottom: 2px !important; }
        .n-b-t { border-top: none !important; padding-top: 2px !important; }
        .p-div { border-bottom: 1px solid rgba(128,128,128,0.06) !important; }
        .bbx-table tbody tr:last-child td { border-bottom: none !important; }
    </style>
    """, unsafe_allow_html=True)
    
    raw_data = []
    bbx_update_text = ""
    if st.session_state.bbx_authenticated:
        raw_data = get_bbx_data_from_sheets()
        if len(raw_data) > 1 and len(raw_data[1]) >= 17:
            bbx_update_text = str(raw_data[1][16]).strip()
            
    col_b_title, col_b_btn = st.columns([3, 1])
    with col_b_title:
        st.markdown("<h1 class='bbx-title'>🛒 BuyBox Takibi</h1>", unsafe_allow_html=True)
        st.markdown("<p class='bbx-subtitle'>Trendyol ve Hepsiburada Buybox durumunuzu takip edin.</p>", unsafe_allow_html=True)
    with col_b_btn:
        if bbx_update_text:
            st.markdown(f'''
                <div class="update-badge">{bbx_update_text}</div>
                <div style="clear: both; height: 10px;"></div>
            ''', unsafe_allow_html=True)
        else:
            st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
            
        _, btn_col = st.columns([1, 1.4]) 
        with btn_col:
            if st.button("🔙 Ana Sayfaya Dön", use_container_width=True):
                st.session_state.current_view = "ana_sayfa"
                st.rerun()
            
    st.markdown("<hr style='margin: 5px 0 15px 0; border: none; border-top: 1px solid rgba(128,128,128,0.2);'>", unsafe_allow_html=True)

    if not st.session_state.bbx_authenticated:
        st.warning("Lütfen ana sayfadan giriş yapınız.")
    else:
        img_map = get_image_mapping()

        if raw_data:
            res_aks = load_and_merge_data()
            df_aks = res_aks[0] if res_aks else None
            ty_retail_map = {}
            hb_retail_map = {}
            
            if df_aks is not None:
                mapping_aks = get_column_mapping(df_aks)
                ty_col = mapping_aks.get("Trendyol")
                hb_col = mapping_aks.get("Hepsiburada")
                
                if "Barkod_Int" in df_aks.columns:
                    for _, r in df_aks.iterrows():
                        b_val = clean_val(str(r["Barkod_Int"]))
                        if b_val:
                            if ty_col and ty_col in df_aks.columns:
                                v = str(r[ty_col]).strip()
                                ty_retail_map[b_val] = v if v.lower() not in ["nan", "none", ""] else "-"
                            if hb_col and hb_col in df_aks.columns:
                                v = str(r[hb_col]).strip()
                                hb_retail_map[b_val] = v if v.lower() not in ["nan", "none", ""] else "-"

            parsed_data = []
            alt_gruplar = set()
            
            for row in raw_data:
                while len(row) < 16: row.append("")
                barkod, sku = str(row[0]).strip(), str(row[2]).strip()
                if "barkod" in barkod.lower() or "fiyat" in barkod.lower() or "satıcı" in barkod.lower(): continue
                if len("".join(row).strip()) < 3: continue
                    
                hb_kod, alt_grup = str(row[1]).strip() if row[1] else "-", str(row[3]).strip() if row[3] else "-"
                barkod, sku = barkod if barkod else "-", sku if sku else "-"
                
                if alt_grup != "-":
                    alt_gruplar.add(alt_grup)

                t1, t2, t3 = str(row[4]).strip().lower(), str(row[6]).strip().lower(), str(row[8]).strip().lower()
                h1, h2, h3 = str(row[10]).strip().lower(), str(row[12]).strip().lower(), str(row[14]).strip().lower()

                def is_ty(sn): return "trendyol" in sn or "braun shop" in sn
                ty_s_txt = ""
                if is_ty(t1): ty_dot, ty_s_txt = "dot-green", "Sorunsuz"
                elif is_ty(t2) or is_ty(t3): ty_dot, ty_s_txt = "dot-yellow", "Fırsat"
                else: ty_dot, ty_s_txt = "dot-red", "Kritik"

                def is_hb(sn): return "hepsiburada" in sn or "braun shop" in sn
                hb_s_txt = ""
                if is_hb(h1): hb_dot, hb_s_txt = "dot-green", "Sorunsuz"
                elif is_hb(h2) or is_hb(h3): hb_dot, hb_s_txt = "dot-yellow", "Fırsat"
                else: hb_dot, hb_s_txt = "dot-red", "Kritik"

                barkod_clean = clean_val(barkod)
                ty_ret = ty_retail_map.get(barkod_clean, "-")
                hb_ret = hb_retail_map.get(barkod_clean, "-")
                
                def ps(v): return f"<span class='pill-satici'>{v}</span>" if v and v != "-" and v.lower() != "nan" else "-"
                def pf(f): return f"<span class='pill-fiyat'>{f} TL</span>" if f and f != "-" and f.lower() != "nan" and "TL" not in f else (f"<span class='pill-fiyat'>{f}</span>" if f and f != "-" and f.lower() != "nan" else "-")
                
                parsed_data.append({
                    "Barkod": barkod, "SKU": sku, "HB Kod": hb_kod, "Alt Grup": alt_grup,
                    "TY Ret": ty_ret, "HB Ret": hb_ret,
                    "TY Durum": ty_s_txt, "TY S1": str(row[4]), "TY F1": str(row[5]), "TY S2": str(row[6]), "TY F2": str(row[7]), "TY S3": str(row[8]), "TY F3": str(row[9]),
                    "HB Durum": hb_s_txt, "HB S1": str(row[10]), "HB F1": str(row[11]), "HB S2": str(row[12]), "HB F2": str(row[13]), "HB S3": str(row[14]), "HB F3": str(row[15]),
                    "_is_ty_alarm": ty_dot in ["dot-yellow", "dot-red"],
                    "_is_hb_alarm": hb_dot in ["dot-yellow", "dot-red"],
                    "_ty_dot": ty_dot, "_hb_dot": hb_dot,
                    "_ts1": ps(row[4]), "_tf1": pf(row[5]), "_ts2": ps(row[6]), "_tf2": pf(row[7]), "_ts3": ps(row[8]), "_tf3": pf(row[9]),
                    "_hs1": ps(row[10]), "_hf1": pf(row[11]), "_hs2": ps(row[12]), "_hf2": pf(row[13]), "_hs3": ps(row[14]), "_hf3": pf(row[15])
                })
                
            col_search, col_grup, col_stat, col_btn_group = st.columns([1.5, 1.5, 1.5, 1.5])
            with col_search: 
                search_bbx = st.text_input("🔍 Ürün Ara...", key="search_bbx")
            with col_grup: 
                grup_bbx = st.multiselect("📂 Alt Grup", sorted(list(alt_gruplar)), placeholder="Tümü", key="grup_bbx")
            with col_stat: 
                alarm_filter = st.selectbox("🚨 Alarm Durumu", ["Tümü", "🔴 Trendyol Alarmları", "🔴 Hepsiburada Alarmları"], key="stat_bbx")
                
            show_ty = alarm_filter in ["Tümü", "🔴 Trendyol Alarmları"]
            show_hb = alarm_filter in ["Tümü", "🔴 Hepsiburada Alarmları"]

            filtered_data = []
            for d in parsed_data:
                if search_bbx:
                    search_str = f"{d['Barkod']} {d['SKU']} {d['HB Kod']}".lower()
                    if search_bbx.lower() not in search_str:
                        continue
                if grup_bbx and d['Alt Grup'] not in grup_bbx:
                    continue
                if alarm_filter == "🔴 Trendyol Alarmları" and not d['_is_ty_alarm']:
                    continue
                if alarm_filter == "🔴 Hepsiburada Alarmları" and not d['_is_hb_alarm']:
                    continue
                
                export_dict = {
                    "Barkod": d["Barkod"], "HB Kod": d["HB Kod"], "SKU": d["SKU"], "Alt Grup": d["Alt Grup"]
                }
                if show_ty:
                    export_dict.update({
                        "Trendyol Retail": d["TY Ret"], "TY Durum": d["TY Durum"], 
                        "TY S1": d["TY S1"], "TY F1": d["TY F1"], 
                        "TY S2": d["TY S2"], "TY F2": d["TY F2"], 
                        "TY S3": d["TY S3"], "TY F3": d["TY F3"]
                    })
                if show_hb:
                    export_dict.update({
                        "Hepsiburada Retail": d["HB Ret"], "HB Durum": d["HB Durum"], 
                        "HB S1": d["HB S1"], "HB F1": d["HB F1"], 
                        "HB S2": d["HB S2"], "HB F2": d["HB F2"], 
                        "HB S3": d["HB S3"], "HB F3": d["HB F3"]
                    })
                
                d["_export_dict"] = export_dict
                filtered_data.append(d)
                
            tr_time_now = datetime.utcnow() + timedelta(hours=3)
            date_str = tr_time_now.strftime("%Y-%m-%d %H:%M:%S")
            
            if alarm_filter == "🔴 Trendyol Alarmları":
                excel_filename = f"TY_BBX_Alarm_{date_str}.xlsx"
            elif alarm_filter == "🔴 Hepsiburada Alarmları":
                excel_filename = f"HB_BBX_Alarm_{date_str}.xlsx"
            else:
                excel_filename = f"Buybox_Takibi_Tümü_{date_str}.xlsx"
                
            with col_btn_group:
                st.markdown("<div style='margin-top: 23px;'></div>", unsafe_allow_html=True)
                btn_clear, btn_excel = st.columns([1, 1])
                with btn_clear:
                    st.button("🧹 Filtre Temizle", on_click=reset_bbx_filters, use_container_width=True, key="btn_clear_bbx")
                with btn_excel:
                    if filtered_data:
                        b_io = io.BytesIO()
                        
                        export_df = pd.DataFrame([d["_export_dict"] for d in filtered_data])
                        
                        price_cols = [c for c in export_df.columns if "Fiyat" in c or "Retail" in c or "F1" in c or "F2" in c or "F3" in c or "Ret" in c]
                        for c in price_cols:
                            export_df[c] = export_df[c].apply(parse_price)
                            
                        if "Barkod" in export_df.columns:
                            def parse_barcode(v):
                                try: return int(float(str(v).replace(' ', '')))
                                except: return v
                            export_df["Barkod"] = export_df["Barkod"].apply(parse_barcode)
                            
                        with pd.ExcelWriter(b_io, engine='openpyxl') as w:
                            export_df.to_excel(w, index=False, sheet_name='BBX_Raporu')
                            worksheet = w.sheets['BBX_Raporu']
                            for idx, col_name in enumerate(export_df.columns):
                                excel_col_idx = idx + 1
                                col_letter = openpyxl.utils.get_column_letter(excel_col_idx)
                                if col_name in price_cols:
                                    worksheet.column_dimensions[col_letter].width = 12
                                    for row in range(2, len(export_df) + 2):
                                        worksheet.cell(row=row, column=excel_col_idx).number_format = '#,##0.00'
                                elif col_name == "Barkod":
                                    worksheet.column_dimensions[col_letter].width = 16
                                    for row in range(2, len(export_df) + 2):
                                        worksheet.cell(row=row, column=excel_col_idx).number_format = '0'
                                else:
                                    worksheet.column_dimensions[col_letter].width = 15
                        
                        st.download_button("📥 Ekrandaki Tabloyu İndir", b_io.getvalue(), excel_filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    else:
                        st.button("📥 Tablo Boş", disabled=True, use_container_width=True, key="empty_bbx")

            ty_l = LOGOS.get("Trendyol", {}).get("light", "")
            ty_d = LOGOS.get("Trendyol", {}).get("dark", "")
            hb_l = LOGOS.get("Hepsiburada", {}).get("light", "")
            hb_d = LOGOS.get("Hepsiburada", {}).get("dark", "")
            
            ty_alarm_count = sum(1 for d in filtered_data if d['_is_ty_alarm'])
            hb_alarm_count = sum(1 for d in filtered_data if d['_is_hb_alarm'])
            
            ty_badge = f"<div style='font-size: 10px; color: #ff1744; font-weight: 700; letter-spacing: 0.5px; text-transform: none; margin-bottom: 3px;'>🚨 {ty_alarm_count} Alarm</div>" if ty_alarm_count > 0 else f"<div style='font-size: 10px; color: #00b84c; font-weight: 700; letter-spacing: 0.5px; text-transform: none; margin-bottom: 3px;'>✨ Kusursuz</div>"
            hb_badge = f"<div style='font-size: 10px; color: #ff1744; font-weight: 700; letter-spacing: 0.5px; text-transform: none; margin-bottom: 3px;'>🚨 {hb_alarm_count} Alarm</div>" if hb_alarm_count > 0 else f"<div style='font-size: 10px; color: #00b84c; font-weight: 700; letter-spacing: 0.5px; text-transform: none; margin-bottom: 3px;'>✨ Kusursuz</div>"

            ty_header = f"""<th colspan="5" style="padding: 6px 0; height: 58px; vertical-align: middle;">{ty_badge}<img src="{ty_l}" class="header-logo logo-light" style="height: 30px;"><img src="{ty_d}" class="header-logo logo-dark" style="height: 30px;"></th>""" if show_ty else ""
            hb_header_class = "plat-sep" if show_ty and show_hb else ""
            hb_header = f"""<th colspan="5" class="{hb_header_class}" style="padding: 6px 0; height: 58px; vertical-align: middle;">{hb_badge}<img src="{hb_l}" class="header-logo logo-light" style="height: 30px;"><img src="{hb_d}" class="header-logo logo-dark" style="height: 30px;"></th>""" if show_hb else ""

            ty_sub = """<th class="shadow-cell" style="width: 85px;">Retail</th><th class="shadow-cell" style="width: 45px;">Durum</th><th class="shadow-cell">#1</th><th class="shadow-cell">#2</th><th class="shadow-cell">#3</th>""" if show_ty else ""
            hb_sub_class = "plat-sep shadow-cell" if show_ty and show_hb else "shadow-cell"
            hb_sub = f"""<th class="{hb_sub_class}" style="width: 85px;">Retail</th><th class="shadow-cell" style="width: 45px;">Durum</th><th class="shadow-cell">#1</th><th class="shadow-cell">#2</th><th class="shadow-cell">#3</th>""" if show_hb else ""

            tbody_html = ""
            for d in filtered_data:
                barkod_clean = clean_val(d["Barkod"])
                img_url = img_map.get(barkod_clean, "")
                sku_cell = f"<td rowspan='2' class='p-div'><div class='sku-wrapper'><span class='bbx-sku'>{d['SKU']}</span><div class='sku-thumb'><img src='{img_url}' referrerpolicy='no-referrer'></div></div></td>" if img_url else f"<td rowspan='2' class='bbx-sku p-div'>{d['SKU']}</td>"
                
                ty_ret_ui = f"{d['TY Ret']} TL" if d['TY Ret'] != "-" else "-"
                hb_ret_ui = f"{d['HB Ret']} TL" if d['HB Ret'] != "-" else "-"
                
                ty_td_class = "p-div"
                ty_cols1 = f"""<td rowspan='2' class='{ty_td_class}' style='font-weight:600; color:var(--text-color); font-size: 11px;'>{ty_ret_ui}</td><td rowspan='2' class='p-div'><div class='status-dot {d['_ty_dot']}'></div></td><td class='n-b-b'>{d['_ts1']}</td><td class='n-b-b'>{d['_ts2']}</td><td class='n-b-b'>{d['_ts3']}</td>""" if show_ty else ""
                ty_cols2 = f"""<td class='n-b-t p-div'>{d['_tf1']}</td><td class='n-b-t p-div'>{d['_tf2']}</td><td class='n-b-t p-div'>{d['_tf3']}</td>""" if show_ty else ""
                
                hb_td_class = "p-div plat-sep" if show_ty and show_hb else "p-div"
                hb_cols1 = f"""<td rowspan='2' class='{hb_td_class}' style='font-weight:600; color:var(--text-color); font-size: 11px;'>{hb_ret_ui}</td><td rowspan='2' class='p-div'><div class='status-dot {d['_hb_dot']}'></div></td><td class='n-b-b'>{d['_hs1']}</td><td class='n-b-b'>{d['_hs2']}</td><td class='n-b-b'>{d['_hs3']}</td>""" if show_hb else ""
                hb_cols2 = f"""<td class='n-b-t p-div'>{d['_hf1']}</td><td class='n-b-t p-div'>{d['_hf2']}</td><td class='n-b-t p-div'>{d['_hf3']}</td>""" if show_hb else ""

                tbody_html += f"<tr><td rowspan='2' class='p-div'>{d['Barkod']}</td><td rowspan='2' class='p-div'>{d['HB Kod']}</td>{sku_cell}<td rowspan='2' class='p-div'>{d['Alt Grup']}</td>{ty_cols1}{hb_cols1}</tr><tr>{ty_cols2}{hb_cols2}</tr>"

            html_table = f"""
            <div class="bbx-table-wrapper">
            <table class="bbx-table">
                <thead>
                    <tr>
                        <th rowspan="2" class="shadow-cell">Barkod</th>
                        <th rowspan="2" class="shadow-cell">HB Kod</th>
                        <th rowspan="2" class="shadow-cell">SKU</th>
                        <th rowspan="2" class="shadow-cell">Alt Grup</th>
                        {ty_header}{hb_header}
                    </tr>
                    <tr>{ty_sub}{hb_sub}</tr>
                </thead>
                <tbody>{tbody_html}</tbody>
            </table></div>
            """
            
            st.markdown(html_table.replace('\n', ''), unsafe_allow_html=True)
